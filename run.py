import streamlit as st
import pandas as pd
import altair as alt
from datetime import datetime, time
import pytz
import io
import calendar
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ============================
# CONFIG TAMPAK APLIKASI
# ============================
st.set_page_config(
    page_title="Aplikasi Akuntansi Keuangan",
    page_icon="💰",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
    /* Pantai Theme */
    .main-title {
        background: linear-gradient(135deg, #56ccf2 0%, #2f80ed 100%);
        padding: 30px;
        border-radius: 15px;
        text-align: center;
        color: white;
        margin-bottom: 30px;
        box-shadow: 0 4px 15px rgba(0,0,0,0.2);
    }
    .main-title h1 {
        font-size: 42px;
        font-weight: 800;
        margin: 0;
        text-shadow: 2px 2px 4px rgba(0,0,0,0.3);
    }
    .main-title p {
        font-size: 18px;
        margin: 10px 0 0 0;
        opacity: 0.9;
    }
    .subtitle {
        background: linear-gradient(135deg, #fbd786 0%, #f7797d 100%);
        padding: 20px;
        border-radius: 10px;
        color: white;
        font-size: 24px;
        font-weight: 700;
        margin: 20px 0;
        text-align: center;
        box-shadow: 0 3px 10px rgba(0,0,0,0.15);
    }
    .metric-card {
        background: white;
        padding: 25px;
        border-radius: 12px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        border-left: 5px solid #2f80ed;
        margin: 10px 0;
    }
    .stButton>button {
        background: linear-gradient(135deg, #56ccf2 0%, #2f80ed 100%) !important;
        color: white !important;
        padding: 12px 28px !important;
        border-radius: 8px !important;
        font-size: 16px !important;
        font-weight: 600 !important;
        border: none !important;
        box-shadow: 0 4px 12px rgba(86,204,242,0.4) !important;
        transition: all 0.3s ease !important;
    }
    .stButton>button:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 6px 20px rgba(86,204,242,0.6) !important;
    }
    /* Sidebar */
    .css-1d391kg, [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #56ccf2 0%, #2f80ed 100%);
    }
    /* Form inputs */
    .stTextInput>div>div>input, .stNumberInput>div>div>input, .stSelectbox>div>div>select {
        border-radius: 8px !important;
        border: 2px solid #e0e0e0 !important;
        padding: 10px !important;
    }
    /* DataFrames */
    .dataframe {
        border-radius: 10px !important;
        overflow: hidden !important;
    }
    /* Info boxes */
    .stAlert {
        border-radius: 10px !important;
        padding: 15px !important;
    }
    /* Metrics */
    [data-testid="stMetricValue"] {
        font-size: 28px !important;
        font-weight: 700 !important;
        color: #2f80ed !important;
    }
    /* Success/Error messages */
    .element-container:has(.stSuccess) {
        animation: slideIn 0.5s ease;
    }
    @keyframes slideIn {
        from {
            opacity: 0;
            transform: translateY(-10px);
        }
        to {
            opacity: 1;
            transform: translateY(0);
        }
    }
</style>
""", unsafe_allow_html=True)

# Header UI
st.markdown("""
<div class='main-title'>
    <h1>💰 Aplikasi Akuntansi Keuangan</h1>
    <p>Kelola keuangan bisnis Anda dengan mudah dan efisien</p>
</div>
""", unsafe_allow_html=True)

# ============================
# SESSION STATE
# ============================
if "transaksi" not in st.session_state:
    st.session_state.transaksi = []

# ============================
# FORMAT RUPIAH Kustom (Rp 4.000.000,00)
# ============================
def to_rp_custom(n):
    try:
        s = f"Rp {int(n):,},00"
        s = s.replace(",", ".")
        return s
    except:
        return "Rp -"

def format_tgl_waktu(dt):
    if pd.isna(dt):
        return ""
    if isinstance(dt, str):
        try:
            dt = pd.to_datetime(dt)
        except:
            return dt
    return dt.strftime("%Y-%m-%d %H:%M:%S")

# ============================
# KLASIFIKASI AKUN
# ============================
pendapatan_akun = ["Pendapatan Jasa", "Pendapatan Lainnya"]
beban_akun = ["Beban Gaji", "Beban Listrik", "Beban Sewa", "Beban Lainnya"]

# ============================
# Fungsi Akuntansi
# ============================
def tambah_transaksi(tgl, akun, ket, debit, kredit):
    st.session_state.transaksi.append({
        "Tanggal": tgl,
        "Akun": akun,
        "Keterangan": ket,
        "Debit": int(debit),
        "Kredit": int(kredit)
    })

def hapus_transaksi(idx):
    st.session_state.transaksi.pop(idx)

def buku_besar(df):
    akun_list = df["Akun"].unique()
    buku_besar_data = {}
    for akun in akun_list:
        df_akun = df[df["Akun"] == akun].copy().sort_values("Tanggal")
        df_akun["Saldo"] = df_akun["Debit"].cumsum() - df_akun["Kredit"].cumsum()
        buku_besar_data[akun] = df_akun
    return buku_besar_data

def neraca_saldo(df):
    grouped = df.groupby("Akun")[["Debit", "Kredit"]].sum()
    grouped["Saldo"] = grouped["Debit"] - grouped["Kredit"]
    return grouped

def laporan_laba_rugi(df):
    total_pendapatan = df[df["Akun"].isin(pendapatan_akun)]["Debit"].sum()
    total_beban = df[df["Akun"].isin(beban_akun)]["Kredit"].sum()
    laba_rugi = total_pendapatan - total_beban
    return {
        "Total Pendapatan": total_pendapatan,
        "Total Beban": total_beban,
        "Laba/Rugi": laba_rugi
    }

# ============================
# EXPORT EXCEL MULTI SHEET SESUAI TEMPLATE
# ============================
def export_excel_multi(df):
    output = io.BytesIO()
    wb = Workbook()
    
    # Styles
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    title_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    year_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    total_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    font_header = Font(color="FFFFFF", bold=True)
    font_title = Font(bold=True)
    
    df["Tanggal"] = pd.to_datetime(df["Tanggal"])
    df = df.sort_values("Tanggal")
    df["Tahun"] = df["Tanggal"].dt.year
    df["Bulan"] = df["Tanggal"].dt.month

    # === SHEET 1: Laporan Keuangan ===
    ws_main = wb.active
    ws_main.title = "Laporan Keuangan"
    current_row = 1
    
    for tahun, df_tahun in df.groupby("Tahun"):
        ws_main.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
        cell = ws_main.cell(row=current_row, column=1, value=f"Laporan Keuangan Tahun {tahun}")
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = year_fill
        for col in range(1, 6):
            ws_main.cell(row=current_row, column=col).border = thin_border
        current_row += 1
        
        for bulan, df_bulan in df_tahun.groupby("Bulan"):
            nama_bulan = calendar.month_name[bulan]
            ws_main.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
            cell_bulan = ws_main.cell(row=current_row, column=1, value=f"Bulan {nama_bulan}")
            cell_bulan.font = Font(bold=True, size=11)
            cell_bulan.alignment = Alignment(horizontal="center", vertical="center")
            cell_bulan.fill = title_fill
            for col in range(1, 6):
                ws_main.cell(row=current_row, column=col).border = thin_border
            current_row += 1

            headers = ["Tanggal", "Akun", "Keterangan", "Debit", "Kredit"]
            for col_num, header in enumerate(headers, start=1):
                cell = ws_main.cell(row=current_row, column=col_num, value=header)
                cell.font = font_header
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = header_fill
                cell.border = thin_border
            current_row += 1

            for _, row in df_bulan.iterrows():
                ws_main.cell(row=current_row, column=1, value=row["Tanggal"].strftime("%Y-%m-%d %H:%M:%S"))
                ws_main.cell(row=current_row, column=1).alignment = Alignment(horizontal="left")
                ws_main.cell(row=current_row, column=1).border = thin_border
                
                ws_main.cell(row=current_row, column=2, value=row["Akun"])
                ws_main.cell(row=current_row, column=2).alignment = Alignment(horizontal="left")
                ws_main.cell(row=current_row, column=2).border = thin_border
                
                ws_main.cell(row=current_row, column=3, value=row["Keterangan"])
                ws_main.cell(row=current_row, column=3).alignment = Alignment(horizontal="left")
                ws_main.cell(row=current_row, column=3).border = thin_border

                debit_cell = ws_main.cell(row=current_row, column=4)
                kredit_cell = ws_main.cell(row=current_row, column=5)

                if row["Debit"] == 0:
                    debit_cell.value = "Rp -"
                else:
                    debit_cell.value = f"Rp {row['Debit']:,},00".replace(",", ".")
                debit_cell.alignment = Alignment(horizontal="right")
                debit_cell.border = thin_border

                if row["Kredit"] == 0:
                    kredit_cell.value = "Rp -"
                else:
                    kredit_cell.value = f"Rp {row['Kredit']:,},00".replace(",", ".")
                kredit_cell.alignment = Alignment(horizontal="right")
                kredit_cell.border = thin_border

                current_row += 1

            # Total per bulan
            total_debit = df_bulan["Debit"].sum()
            total_kredit = df_bulan["Kredit"].sum()

            ws_main.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
            cell_total = ws_main.cell(row=current_row, column=1, value="TOTAL")
            cell_total.font = Font(bold=True)
            cell_total.alignment = Alignment(horizontal="center")
            cell_total.fill = total_fill
            for col in range(1, 6):
                ws_main.cell(row=current_row, column=col).border = thin_border

            d_cell = ws_main.cell(row=current_row, column=4)
            k_cell = ws_main.cell(row=current_row, column=5)
            if total_debit == 0:
                d_cell.value = "Rp -"
            else:
                d_cell.value = f"Rp {total_debit:,},00".replace(",", ".")
            d_cell.font = Font(bold=True)
            d_cell.alignment = Alignment(horizontal="right")
            d_cell.fill = total_fill
            d_cell.border = thin_border

            if total_kredit == 0:
                k_cell.value = "Rp -"
            else:
                k_cell.value = f"Rp {total_kredit:,},00".replace(",", ".")
            k_cell.font = Font(bold=True)
            k_cell.alignment = Alignment(horizontal="right")
            k_cell.fill = total_fill
            k_cell.border = thin_border

            current_row += 2

    for col, width in zip(['A', 'B', 'C', 'D', 'E'], [22, 18, 22, 22, 22]):
        ws_main.column_dimensions[col].width = width

    # === SHEET 2: Jurnal Umum ===
    ws_jurnal = wb.create_sheet("Jurnal Umum")
    ws_jurnal.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    c = ws_jurnal.cell(row=1, column=1, value="Jurnal Umum")
    c.font = Font(bold=True, size=14)
    c.alignment = Alignment(horizontal="center")

    headers = ["Tanggal", "Akun", "Keterangan", "Debit", "Kredit"]
    for col_num, header in enumerate(headers, start=1):
        c = ws_jurnal.cell(row=2, column=col_num, value=header)
        c.font = font_header
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = header_fill
        c.border = thin_border

    row_num = 3
    for _, row in df.iterrows():
        ws_jurnal.cell(row=row_num, column=1, value=row["Tanggal"].strftime("%Y-%m-%d %H:%M:%S")).alignment = Alignment(horizontal="left")
        ws_jurnal.cell(row=row_num, column=1).border = thin_border
        ws_jurnal.cell(row=row_num, column=2, value=row["Akun"]).alignment = Alignment(horizontal="left")
        ws_jurnal.cell(row=row_num, column=2).border = thin_border
        ws_jurnal.cell(row=row_num, column=3, value=row["Keterangan"]).alignment = Alignment(horizontal="left")
        ws_jurnal.cell(row=row_num, column=3).border = thin_border

        d_cell = ws_jurnal.cell(row=row_num, column=4)
        k_cell = ws_jurnal.cell(row=row_num, column=5)

        if row["Debit"] == 0:
            d_cell.value = "Rp -"
        else:
            d_cell.value = f"Rp {row['Debit']:,},00".replace(",", ".")
        d_cell.alignment = Alignment(horizontal="right")
        d_cell.border = thin_border

        if row["Kredit"] == 0:
            k_cell.value = "Rp -"
        else:
            k_cell.value = f"Rp {row['Kredit']:,},00".replace(",", ".")
        k_cell.alignment = Alignment(horizontal="right")
        k_cell.border = thin_border

        row_num += 1

    for col, width in zip(['A', 'B', 'C', 'D', 'E'], [22, 18, 22, 22, 22]):
        ws_jurnal.column_dimensions[col].width = width

    # === SHEET 3: Buku Besar ===
    ws_bb = wb.create_sheet("Buku Besar")
    bb = buku_besar(df)
    curr_row_bb = 1

    for akun, data in bb.items():
        ws_bb.merge_cells(start_row=curr_row_bb, start_column=1, end_row=curr_row_bb, end_column=6)
        c = ws_bb.cell(row=curr_row_bb, column=1, value=f"Buku Besar - {akun}")
        c.font = Font(bold=True, size=12)
        c.alignment = Alignment(horizontal="center")
        curr_row_bb += 1

        headers = ["Tanggal", "Akun", "Keterangan", "Debit", "Kredit", "Saldo"]
        for col_num, header in enumerate(headers, start=1):
            c = ws_bb.cell(row=curr_row_bb, column=col_num, value=header)
            c.font = font_header
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.fill = header_fill
            c.border = thin_border
        curr_row_bb += 1

        for _, row in data.iterrows():
            ws_bb.cell(row=curr_row_bb, column=1, value=row["Tanggal"].strftime("%Y-%m-%d %H:%M:%S")).alignment = Alignment(horizontal="left")
            ws_bb.cell(row=curr_row_bb, column=1).border = thin_border

            ws_bb.cell(row=curr_row_bb, column=2, value=row["Akun"]).alignment = Alignment(horizontal="left")
            ws_bb.cell(row=curr_row_bb, column=2).border = thin_border

            ws_bb.cell(row=curr_row_bb, column=3, value=row["Keterangan"]).alignment = Alignment(horizontal="left")
            ws_bb.cell(row=curr_row_bb, column=3).border = thin_border

            debit = row["Debit"]
            kredit = row["Kredit"]
            saldo = row["Saldo"]

            d_cell = ws_bb.cell(row=curr_row_bb, column=4)
            k_cell = ws_bb.cell(row=curr_row_bb, column=5)
            s_cell = ws_bb.cell(row=curr_row_bb, column=6)

            d_cell.value = f"Rp {debit:,},00".replace(",", ".") if debit != 0 else "Rp -"
            d_cell.alignment = Alignment(horizontal="right")
            d_cell.border = thin_border

            k_cell.value = f"Rp {kredit:,},00".replace(",", ".") if kredit != 0 else "Rp -"
            k_cell.alignment = Alignment(horizontal="right")
            k_cell.border = thin_border

            s_cell.value = f"Rp {saldo:,},00".replace(",", ".") if saldo != 0 else "Rp -"
            s_cell.alignment = Alignment(horizontal="right")
            s_cell.border = thin_border

            curr_row_bb += 1

        curr_row_bb += 2

    for col, width in zip(['A', 'B', 'C', 'D', 'E', 'F'], [22, 18, 22, 22, 22, 22]):
        ws_bb.column_dimensions[col].width = width

    # === SHEET 4: Neraca Saldo ===
    ws_ns = wb.create_sheet("Neraca Saldo")
    ns = neraca_saldo(df)
    ws_ns.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    c = ws_ns.cell(row=1, column=1, value="Neraca Saldo")
    c.font = Font(bold=True, size=14)
    c.alignment = Alignment(horizontal="center")

    headers = ["Akun", "Debit", "Kredit", "Saldo"]
    for col_num, header in enumerate(headers, start=1):
        c = ws_ns.cell(row=2, column=col_num, value=header)
        c.font = font_header
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = header_fill
        c.border = thin_border

    row_num = 3
    for _, row in ns.reset_index().iterrows():
        ws_ns.cell(row=row_num, column=1, value=row["Akun"]).alignment = Alignment(horizontal="left")
        ws_ns.cell(row=row_num, column=1).border = thin_border

        d_cell = ws_ns.cell(row=row_num, column=2)
        k_cell = ws_ns.cell(row=row_num, column=3)
        s_cell = ws_ns.cell(row=row_num, column=4)

        d = row["Debit"]
        k = row["Kredit"]
        s = row["Saldo"]

        d_cell.value = f"Rp {d:,},00".replace(",", ".") if d != 0 else "Rp -"
        d_cell.alignment = Alignment(horizontal="right")
        d_cell.border = thin_border

        k_cell.value = f"Rp {k:,},00".replace(",", ".") if k != 0 else "Rp -"
        k_cell.alignment = Alignment(horizontal="right")
        k_cell.border = thin_border

        s_cell.value = f"Rp {s:,},00".replace(",", ".") if s != 0 else "Rp -"
        s_cell.alignment = Alignment(horizontal="right")
        s_cell.border = thin_border

        row_num += 1

    for col, width in zip(['A', 'B', 'C', 'D'], [22, 22, 22, 22]):
        ws_ns.column_dimensions[col].width = width

    # === SHEET 5: Laporan Laba Rugi ===
    ws_lr = wb.create_sheet("Laporan Laba Rugi")
    lr = laporan_laba_rugi(df)

    ws_lr.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    c = ws_lr.cell(row=1, column=1, value="Laporan Laba Rugi")
    c.font = Font(bold=True, size=14)
    c.alignment = Alignment(horizontal="center")
    c.fill = year_fill

    headers = ["Keterangan", "Jumlah"]
    for col_num, header in enumerate(headers, start=1):
        c = ws_lr.cell(row=2, column=col_num, value=header)
        c.font = font_header
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = header_fill
        c.border = thin_border

    rows = [
        ("Total Pendapatan", lr["Total Pendapatan"]),
        ("Total Beban", lr["Total Beban"]),
        ("Laba/Rugi", lr["Laba/Rugi"])
    ]
    current_row = 3
    for keterangan, jumlah in rows:
        c1 = ws_lr.cell(row=current_row, column=1, value=keterangan)
        c1.alignment = Alignment(horizontal="left")
        c1.border = thin_border

        c2 = ws_lr.cell(row=current_row, column=2)
        c2.alignment = Alignment(horizontal="right")
        c2.border = thin_border

        if keterangan == "Laba/Rugi" and jumlah < 0:
            c2.value = f"(Rp {abs(jumlah):,},00)".replace(",", ".")
        elif jumlah == 0:
            c2.value = "Rp -"
        else:
            c2.value = f"Rp {jumlah:,},00".replace(",", ".")
        current_row += 1

    ws_lr.column_dimensions['A'].width = 22
    ws_lr.column_dimensions['B'].width = 22

    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ============================
# MENU NAVIGASI
# ============================
st.sidebar.markdown("### 📋 Menu Navigasi")
menu = st.sidebar.radio(
    "",
    ["🏠 Dashboard", "📝 Input Transaksi", "📋 Lihat Transaksi", "📖 Buku Besar",
     "⚖️ Neraca Saldo", "💰 Laporan Laba Rugi", "📈 Grafik", "📥 Import Excel", "📤 Export Excel"],
    label_visibility="collapsed"
)

st.sidebar.markdown("---")
st.sidebar.markdown("### 📊 Statistik")
total_transaksi = len(st.session_state.transaksi)
st.sidebar.info(f"Total Transaksi: **{total_transaksi}**")

if total_transaksi > 0:
    df_temp = pd.DataFrame(st.session_state.transaksi)
    total_debit = df_temp["Debit"].sum()
    total_kredit = df_temp["Kredit"].sum()
    st.sidebar.success(f"Total Debit: **Rp {total_debit:,},00**".replace(",", "."))
    st.sidebar.warning(f"Total Kredit: **Rp {total_kredit:,},00**".replace(",", "."))

# ============================
# 0. DASHBOARD
# ============================
if menu == "🏠 Dashboard":
    st.markdown("<div class='subtitle'>🏠 Dashboard Overview</div>", unsafe_allow_html=True)

    if len(st.session_state.transaksi) == 0:
        st.info("👋 Selamat datang! Mulai dengan menambahkan transaksi pertama Anda.")
        st.markdown("""
### 📚 Panduan Penggunaan:
1. **Input Transaksi** - Tambahkan transaksi baru
2. **Lihat Transaksi** - Review dan hapus transaksi
3. **Buku Besar** - Lihat detail per akun
4. **Neraca Saldo** - Ringkasan semua akun
5. **Laporan Laba Rugi** - Analisis profit/loss
6. **Grafik** - Visualisasi data
7. **Import Excel** - Import transaksi dari Excel
8. **Export Excel** - Download laporan lengkap
        """)
    else:
        df = pd.DataFrame(st.session_state.transaksi)
        lr = laporan_laba_rugi(df)

        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("📊 Total Transaksi", f"{len(df)}")
        with col2:
            st.metric("💵 Total Pendapatan", to_rp_custom(lr["Total Pendapatan"]))
        with col3:
            st.metric("💸 Total Beban", to_rp_custom(lr["Total Beban"]))
        with col4:
            if lr["Laba/Rugi"] >= 0:
                st.metric("✅ Laba Bersih", to_rp_custom(lr["Laba/Rugi"]))
            else:
                st.metric("⚠️ Rugi Bersih", to_rp_custom(abs(lr["Laba/Rugi"])))
        st.markdown("---")

        st.markdown("### 📋 Transaksi Terbaru")
        df_display = df.tail(5).copy()
        df_display["Debit"] = df_display["Debit"].apply(to_rp_custom)
        df_display["Kredit"] = df_display["Kredit"].apply(to_rp_custom)
        df_display["Tanggal"] = df_display["Tanggal"].apply(format_tgl_waktu)
        st.dataframe(df_display, use_container_width=True, hide_index=True)

# ============================
# 1. INPUT TRANSAKSI DENGAN WAKTU ASIA/JAKARTA OTOMATIS
# ============================
elif menu == "📝 Input Transaksi":
    st.markdown("<div class='subtitle'>📝 Input Transaksi Baru</div>", unsafe_allow_html=True)

    st.markdown("""
<div style='background: #e3f2fd; padding: 15px; border-radius: 10px; margin-bottom: 20px;'>
    <h4 style='color: #1976d2; margin: 0;'>💡 Tips Pencatatan:</h4>
    <ul style='color: #1976d2; margin: 5px 0;'>
        <li><strong>Pendapatan</strong> → Dicatat di kolom <strong>DEBIT</strong></li>
        <li><strong>Beban</strong> → Dicatat di kolom <strong>KREDIT</strong></li>
        <li><strong>Kas Masuk</strong> → Kas di <strong>DEBIT</strong>, Pendapatan di <strong>DEBIT</strong></li>
        <li><strong>Kas Keluar</strong> → Beban di <strong>KREDIT</strong>, Kas di <strong>KREDIT</strong></li>
    </ul>
</div>
    """, unsafe_allow_html=True)

    with st.form("form_transaksi", clear_on_submit=True):
        col1, col2 = st.columns(2)

        with col1:
            # Waktu Asia Jakarta
            tz = pytz.timezone('Asia/Jakarta')
            tgl_input = st.date_input("📅 Tanggal Transaksi", datetime.now(tz).date())
            akun = st.selectbox("🏦 Pilih Akun",
                                ["Kas", "Piutang", "Modal", "Pendapatan Jasa", "Pendapatan Lainnya",
                                 "Beban Gaji", "Beban Listrik", "Beban Sewa", "Beban Lainnya"])
            ket = st.text_input("📝 Keterangan", placeholder="Contoh: Pembayaran gaji karyawan")

        with col2:
            st.markdown("#### 💰 Jumlah Transaksi")
            debit = st.number_input("Debit (Rp)", min_value=0, step=10000, format="%d")
            kredit = st.number_input("Kredit (Rp)", min_value=0, step=10000, format="%d")

        st.markdown("---")
        col_btn1, col_btn2 = st.columns([1,1])
        with col_btn1:
            submit = st.form_submit_button("✅ Simpan Transaksi", use_container_width=True)
        with col_btn2:
            reset = st.form_submit_button("🔄 Reset")

        if submit:
            if debit == 0 and kredit == 0:
                st.error("❌ Debit atau Kredit harus diisi!")
            elif ket.strip() == "":
                st.error("❌ Keterangan harus diisi!")
            else:
                waktu_device = datetime.now(tz).time()
                tgl_waktu = datetime.combine(tgl_input, waktu_device)
                tambah_transaksi(tgl_waktu, akun, ket, debit, kredit)
                st.success("✅ Transaksi berhasil ditambahkan!")
                st.balloons()
                st.experimental_rerun()

# ============================
# 2. LIHAT TRANSAKSI
# ============================
elif menu == "📋 Lihat Transaksi":
    st.markdown("<div class='subtitle'>📋 Daftar Semua Transaksi</div>", unsafe_allow_html=True)

    if len(st.session_state.transaksi) == 0:
        st.info("📭 Belum ada transaksi yang tercatat.")
    else:
        df = pd.DataFrame(st.session_state.transaksi)

        col_f1, col_f2 = st.columns(2)
        with col_f1:
            filter_akun = st.multiselect("🔍 Filter berdasarkan Akun", df["Akun"].unique())
        with col_f2:
            sort_by = st.selectbox("📊 Urutkan berdasarkan", ["Tanggal", "Akun", "Debit", "Kredit"])

        df_filtered = df.copy()
        if filter_akun:
            df_filtered = df_filtered[df_filtered["Akun"].isin(filter_akun)]
        df_filtered = df_filtered.sort_values(sort_by)

        df_display = df_filtered.copy()
        df_display["Debit"] = df_display["Debit"].apply(to_rp_custom)
        df_display["Kredit"] = df_display["Kredit"].apply(to_rp_custom)
        df_display["Tanggal"] = df_display["Tanggal"].apply(format_tgl_waktu)

        st.dataframe(df_display, use_container_width=True, hide_index=True)

        st.markdown("---")
        st.markdown("### 🗑️ Hapus Transaksi")
        col_h1, col_h2 = st.columns([3, 1])
        with col_h1:
            idx_hapus = st.number_input("Nomor indeks transaksi yang ingin dihapus",
                                       min_value=0, max_value=len(st.session_state.transaksi)-1, step=1)
        with col_h2:
            if st.button("🗑️ Hapus", use_container_width=True):
                hapus_transaksi(idx_hapus)
                st.success("✅ Transaksi berhasil dihapus!")
                st.experimental_rerun()

# ============================
# 3. BUKU BESAR
# ============================
elif menu == "📖 Buku Besar":
    st.markdown("<div class='subtitle'>📖 Buku Besar Per Akun</div>", unsafe_allow_html=True)

    if len(st.session_state.transaksi) == 0:
        st.info("📭 Belum ada transaksi untuk ditampilkan.")
    else:
        df = pd.DataFrame(st.session_state.transaksi)
        bb = buku_besar(df)

        for idx, (akun, data) in enumerate(bb.items()):
            with st.expander(f"📊 {akun}", expanded=(idx == 0)):
                data_display = data.copy()
                data_display["Debit"] = data_display["Debit"].apply(to_rp_custom)
                data_display["Kredit"] = data_display["Kredit"].apply(to_rp_custom)
                data_display["Saldo"] = data_display["Saldo"].apply(to_rp_custom)
                data_display["Tanggal"] = data_display["Tanggal"].apply(format_tgl_waktu)
                st.dataframe(data_display, use_container_width=True, hide_index=True)

# ============================
# 4. NERACA SALDO
# ============================
elif menu == "⚖️ Neraca Saldo":
    st.markdown("<div class='subtitle'>⚖️ Neraca Saldo</div>", unsafe_allow_html=True)

    if len(st.session_state.transaksi) == 0:
        st.info("📭 Belum ada transaksi untuk ditampilkan.")
    else:
        df = pd.DataFrame(st.session_state.transaksi)
        ns = neraca_saldo(df)
        ns_display = ns.copy()
        ns_display["Debit"] = ns_display["Debit"].apply(to_rp_custom)
        ns_display["Kredit"] = ns_display["Kredit"].apply(to_rp_custom)
        ns_display["Saldo"] = ns_display["Saldo"].apply(to_rp_custom)
        st.dataframe(ns_display, use_container_width=True)

# ============================
# 5. LAPORAN LABA RUGI
# ============================
elif menu == "💰 Laporan Laba Rugi":
    st.markdown("<div class='subtitle'>💰 Laporan Laba Rugi</div>", unsafe_allow_html=True)

    if len(st.session_state.transaksi) == 0:
        st.info("📭 Belum ada transaksi untuk dianalisis.")
    else:
        df = pd.DataFrame(st.session_state.transaksi)
        lr = laporan_laba_rugi(df)

        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown(f"""
            <div style='background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%);
                        padding: 25px; border-radius: 12px; color: white; text-align: center;'>
                <h3 style='margin: 0; font-size: 18px;'>💵 Total Pendapatan</h3>
                <h2 style='margin: 10px 0 0 0; font-size: 28px;'>{to_rp_custom(lr["Total Pendapatan"])}</h2>
            </div>
            """, unsafe_allow_html=True)
        with col2:
            st.markdown(f"""
            <div style='background: linear-gradient(135deg, #ee0979 0%, #ff6a00 100%);
                        padding: 25px; border-radius: 12px; color: white; text-align: center;'>
                <h3 style='margin: 0; font-size: 18px;'>💸 Total Beban</h3>
                <h2 style='margin: 10px 0 0 0; font-size: 28px;'>{to_rp_custom(lr["Total Beban"])}</h2>
            </div>
            """, unsafe_allow_html=True)
        with col3:
            if lr["Laba/Rugi"] >= 0:
                st.markdown(f"""
                <div style='background: linear-gradient(135deg, #56ccf2 0%, #2f80ed 100%);
                            padding: 25px; border-radius: 12px; color: white; text-align: center;'>
                    <h3 style='margin: 0; font-size: 18px;'>✅ Laba Bersih</h3>
                    <h2 style='margin: 10px 0 0 0; font-size: 28px;'>{to_rp_custom(lr["Laba/Rugi"])}</h2>
                </div>
                """, unsafe_allow_html=True)
            else:
                st.markdown(f"""
                <div style='background: linear-gradient(135deg, #f2709c 0%, #ff9472 100%);
                            padding: 25px; border-radius: 12px; color: white; text-align: center;'>
                    <h3 style='margin: 0; font-size: 18px;'>⚠️ Rugi Bersih</h3>
                    <h2 style='margin: 10px 0 0 0; font-size: 28px;'>{to_rp_custom(abs(lr["Laba/Rugi"]))}</h2>
                </div>
                """, unsafe_allow_html=True)

        st.markdown("---")
        st.markdown("### 📊 Detail Perhitungan")
        detail_data = {
            "Keterangan": ["Total Pendapatan", "Total Beban", "Laba/Rugi Bersih"],
            "Jumlah": [to_rp_custom(lr["Total Pendapatan"]), to_rp_custom(lr["Total Beban"]),
                      to_rp_custom(lr["Laba/Rugi"]) if lr["Laba/Rugi"] >= 0 else f"({to_rp_custom(abs(lr['Laba/Rugi']))})"]
        }
        st.table(pd.DataFrame(detail_data))

# ============================
# 6. GRAFIK
# ============================
elif menu == "📈 Grafik":
    st.markdown("<div class='subtitle'>📈 Visualisasi Data Akuntansi</div>", unsafe_allow_html=True)

    if len(st.session_state.transaksi) == 0:
        st.info("📭 Belum ada data untuk divisualisasikan.")
    else:
        df = pd.DataFrame(st.session_state.transaksi)

        tab1, tab2, tab3 = st.tabs(["📊 Debit per Akun", "📊 Kredit per Akun", "📊 Perbandingan"])

        with tab1:
            chart = alt.Chart(df).mark_bar().encode(
                x=alt.X("Akun:N", title="Akun"),
                y=alt.Y("Debit:Q", title="Debit (Rp)"),
                color=alt.Color("Akun:N", legend=None),
                tooltip=["Akun", "Debit"]
            ).properties(title="Grafik Total Debit per Akun", height=400)
            st.altair_chart(chart, use_container_width=True)

        with tab2:
            chart2 = alt.Chart(df).mark_bar().encode(
                x=alt.X("Akun:N", title="Akun"),
                y=alt.Y("Kredit:Q", title="Kredit (Rp)"),
                color=alt.Color("Akun:N", legend=None),
                tooltip=["Akun", "Kredit"]
            ).properties(title="Grafik Total Kredit per Akun", height=400)
            st.altair_chart(chart2, use_container_width=True)

        with tab3:
            df_grouped = df.groupby("Akun")[["Debit", "Kredit"]].sum().reset_index()
            df_melted = df_grouped.melt(id_vars="Akun", value_vars=["Debit", "Kredit"],
                                        var_name="Tipe", value_name="Jumlah")

            chart3 = alt.Chart(df_melted).mark_bar().encode(
                x=alt.X("Akun:N", title="Akun"),
                y=alt.Y("Jumlah:Q", title="Jumlah (Rp)"),
                color="Tipe:N",
                xOffset="Tipe:N",
                tooltip=["Akun", "Tipe", "Jumlah"]
            ).properties(title="Perbandingan Debit vs Kredit per Akun", height=400)
            st.altair_chart(chart3, use_container_width=True)

# ============================
# 7. IMPORT EXCEL
# ============================
elif menu == "📥 Import Excel":
    st.markdown("<div class='subtitle'>📥 Import Transaksi dari File Excel</div>", unsafe_allow_html=True)

    uploaded_file = st.file_uploader("Upload file Excel (.xlsx)", type=["xlsx"])

    if uploaded_file:
        try:
            df_import = pd.read_excel(uploaded_file)
            expected_cols = ["Tanggal", "Akun", "Keterangan", "Debit", "Kredit"]
            if all(col in df_import.columns for col in expected_cols):
                df_import["Tanggal"] = pd.to_datetime(df_import["Tanggal"], errors='coerce')
                df_import = df_import.dropna(subset=["Tanggal"])

                df_import["Debit"] = pd.to_numeric(df_import["Debit"], errors='coerce').fillna(0).astype(int)
                df_import["Kredit"] = pd.to_numeric(df_import["Kredit"], errors='coerce').fillna(0).astype(int)

                st.markdown("### Preview Data yang akan diimpor")
                preview = df_import.copy()
                preview["Tanggal"] = preview["Tanggal"].apply(format_tgl_waktu)
                preview["Debit"] = preview["Debit"].apply(to_rp_custom)
                preview["Kredit"] = preview["Kredit"].apply(to_rp_custom)
                st.dataframe(preview)

                if st.button("🗂️ Tambahkan Semua Transaksi dari File"):
                    for idx, row in df_import.iterrows():
                        tambah_transaksi(row["Tanggal"], row["Akun"], row["Keterangan"], row["Debit"], row["Kredit"])
                    st.success(f"✅ Berhasil menambahkan {len(df_import)} transaksi!")
                    st.experimental_rerun()
            else:
                st.error(f"Format file Excel tidak sesuai. Pastikan kolom: {expected_cols} ada.")
        except Exception as e:
            st.error(f"Gagal membaca file Excel: {e}")

# ============================
# 8. EXPORT EXCEL
# ============================
elif menu == "📤 Export Excel":
    st.markdown("<div class='subtitle'>📤 Export Laporan ke Excel</div>", unsafe_allow_html=True)

    if len(st.session_state.transaksi) == 0:
        st.info("📭 Belum ada transaksi untuk diekspor.")
    else:
        st.markdown("""
<div style='background: #fff3cd; padding: 20px; border-radius: 10px; border-left: 5px solid #ffc107; margin-bottom: 20px;'>
    <h4 style='color: #856404; margin: 0 0 10px 0;'>📦 File Excel akan berisi:</h4>
    <ul style='color: #856404; margin: 0;'>
        <li>📄 Sheet 1: Laporan Keuangan (per bulan dan tahun)</li>
        <li>📄 Sheet 2: Jurnal Umum</li>
        <li>📄 Sheet 3: Buku Besar</li>
        <li>📄 Sheet 4: Neraca Saldo</li>
        <li>📄 Sheet 5: Laporan Laba Rugi</li>
    </ul>
</div>
        """, unsafe_allow_html=True)

        df = pd.DataFrame(st.session_state.transaksi)

        st.markdown("### 👁️ Preview Data")
        col_p1, col_p2 = st.columns(2)
        with col_p1:
            st.metric("Total Transaksi", len(df))
        with col_p2:
            st.metric("Total Akun Unik", df["Akun"].nunique())

        try:
            excel_file = export_excel_multi(df)

            st.markdown("---")
            col_d1, col_d2, col_d3 = st.columns([1, 2, 1])
            with col_d2:
                st.download_button(
                    label="📥 Download Laporan Excel Lengkap",
                    data=excel_file,
                    file_name=f"laporan_akuntansi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            st.success("✅ File Excel siap didownload!")
        except Exception as e:
            st.error(f"❌ Terjadi kesalahan saat membuat file Excel: {str(e)}")

# Footer
st.markdown("---")
st.markdown("""
<div style='text-align: center; color: #888; padding: 20px;'>
    <p style='margin: 0;'>💰 <strong>Aplikasi Akuntansi Profesional</strong></p>
    <p style='margin: 5px 0 0 0; font-size: 14px;'>Kelola keuangan bisnis Anda dengan mudah dan efisien</p>
</div>
""", unsafe_allow_html=True)
